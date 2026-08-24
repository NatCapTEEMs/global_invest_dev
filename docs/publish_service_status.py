#!/usr/bin/env python3
"""Publish docs/service_status.qmd to the group's shared status sheet.

The qmd is the source of truth and is written to be edited by hand, the same way the deck is:
one `## Service` section per service, one `- **field:** value` bullet per field. This reads
those sections and writes the sheet, so the prose lives in one place under version control and
a status change shows up in a diff beside the code change that caused it.

The `method` field is the one-line answer to "how is this computed", written as the formula in
words. The long version is each module's `<service>_method.qmd`, which is where a reader who
wants the equations should go; this column exists so the sheet can be read on its own.

The `ours` field is the answer to whether we can produce the number independently, which is not
the same question as whether a run completes. It takes three values. `yes` means we compute it
here from inputs. `model here` means the thing that produced it is installed and we take its
output as given anyway, so this is a choice rather than a limit. `no model` means what produced
the number is not something we hold, so the number cannot be ours whatever we do.

    python3 docs/publish_service_status.py            # publish
    python3 docs/publish_service_status.py --check    # parse and report, write nothing

It writes to the SAME sheet every time, so the group's link keeps working. Row heights are set
from the content because Google keeps a one-line height on rows it has seen before, and a
wrapped cell then renders clipped however the wrap flag is set.
"""
import math
import os
import re
import subprocess
import sys
import tempfile

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

QMD_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'service_status.qmd')
SHEET_REMOTE = 'gdrive:gep/gep_library_status.xlsx'
SHEET_TITLE = 'library status'

# The bullet label in the qmd -> the sheet's column heading, in the sheet's column order.
FIELD_COLUMNS = [
    ('subgroup', 'subgroup'),
    ('module', 'module'),
    ('category', 'what the code does'),
    ('in the library', 'in the library'),
    ('run', 'run'),
    ('ours', 'can we run it ourselves'),
    ('method', 'how it is computed'),
    ('code', 'what we did to the code'),
    ('total', 'total (2019 USD)'),
    ('number', 'what the number is, and what it is checked against'),
    ('need', 'what we need'),
]
COLUMN_WIDTHS = (22, 16, 22, 26, 11, 8, 30, 52, 64, 16, 64, 40)
LINE_HEIGHT_POINTS = 13.5
MAX_LINES = 14

SECTION_PATTERN = re.compile(r'^## (.+?)\s*$', re.M)
BULLET_PATTERN = re.compile(r'^- \*\*(.+?):\*\*\s*(.*)$', re.M)


def parse_qmd(path):
    """The qmd's service sections as one row per service.

    Raises:
        ValueError: if a section is missing a field, or carries one the sheet has no column
            for. Publishing a table with a silently empty column is how the sheet and the
            prose drift apart.
    """
    text = open(path, encoding='utf-8').read()
    body = text[text.index('\n---', 3) + 4:] if text.startswith('---') else text

    sections = list(SECTION_PATTERN.finditer(body))
    known = {label for label, _ in FIELD_COLUMNS}
    rows, problems = [], []
    for index, match in enumerate(sections):
        name = match.group(1).strip()
        end = sections[index + 1].start() if index + 1 < len(sections) else len(body)
        chunk = body[match.end():end]
        fields = {m.group(1).strip(): m.group(2).strip() for m in BULLET_PATTERN.finditer(chunk)}

        unknown = set(fields) - known
        if unknown:
            problems.append(f'{name}: unknown field(s) {sorted(unknown)}')
        missing = known - set(fields)
        if missing:
            problems.append(f'{name}: missing field(s) {sorted(missing)}')

        row = {'service': name}
        for label, column in FIELD_COLUMNS:
            row[column] = fields.get(label, '')
        rows.append(row)

    if problems:
        raise ValueError('service_status.qmd does not match the sheet\'s columns:\n  '
                         + '\n  '.join(problems))
    if not rows:
        raise ValueError(f'no service sections found in {path}')
    columns = ['service'] + [column for _, column in FIELD_COLUMNS]
    return pd.DataFrame(rows)[columns]


def build_workbook(df):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = SHEET_TITLE
    sheet.append(list(df.columns))
    for record in df.to_dict('records'):
        sheet.append([record[column] for column in df.columns])

    for index, width in enumerate(COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    # The header scrolls out of view within three services otherwise, and the columns are long
    # enough that nobody can tell which one they are reading without it.
    sheet.freeze_panes = 'B2'

    for row_index in range(1, sheet.max_row + 1):
        lines = 1
        for column_index, width in enumerate(COLUMN_WIDTHS, start=1):
            text = str(sheet.cell(row_index, column_index).value or '')
            if text:
                lines = max(lines, math.ceil(len(text) / width))
        sheet.row_dimensions[row_index].height = round(
            min(lines, MAX_LINES) * LINE_HEIGHT_POINTS + 4, 1)
        for column_index in range(1, len(COLUMN_WIDTHS) + 1):
            sheet.cell(row_index, column_index).alignment = Alignment(
                wrap_text=True, vertical='top')
    return workbook


def main():
    df = parse_qmd(QMD_PATH)
    if '--check' in sys.argv:
        print(f'{len(df)} services parse cleanly from {os.path.basename(QMD_PATH)}')
        print(df[['service', 'module', 'total (2019 USD)']].to_string(index=False))
        return

    with tempfile.TemporaryDirectory() as directory:
        path = os.path.join(directory, 'gep_library_status.xlsx')
        build_workbook(df).save(path)
        result = subprocess.run(
            ['rclone', 'copyto', path, SHEET_REMOTE, '--drive-import-formats', 'xlsx'],
            capture_output=True, text=True)
    if result.returncode != 0:
        print(result.stderr, file=sys.stderr)
        raise SystemExit(f'rclone failed with code {result.returncode}')
    print(f'published {len(df)} services to {SHEET_REMOTE}')


if __name__ == '__main__':
    main()
